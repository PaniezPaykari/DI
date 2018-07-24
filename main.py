#import requests
#url = 'http://www1.nyc.gov/assets/ccrb/downloads/excel/ccrb_datatransparencyinitiative_20170207.xlsx'

import openpyxl as px
import numpy as np
import pandas as pd
from collections import Counter


workbook  = px.load_workbook('ccrb_datatransparencyinitiative_20170207.xlsx')
sheet     = workbook['Complaints_Allegations']
row_count = sheet.max_row
col_count = sheet.max_column

data_arr = []
for r in sheet.iter_rows():
    for rr in r:
        data_arr.append(rr.internal_value)

print(len(data_arr)) #3307504
print(row_count*col_count) # 3307504

data         = np.resize(data_arr,[row_count,col_count])
data_as_list = data.tolist()

ID_index = data_as_list[0].index('UniqueComplaintId')
ID_count = len(Counter(data[1:,ID_index]))

NAID_count = []
for x in range(row_count):
    for y in range(col_count):
        if sheet.cell(x+1,y+1).value == 'NA':
            NAID_count.append(sheet.cell(x+1,2).value)

IncompleteInfo_count = len(Counter(NAID_count))
CompleteInfo_count   = ID_count - IncompleteInfo_count

--------------------------------------------------------------

data_completeInfo = data.copy()
for i in range(len(NAID_count)):
    NAID_index = np.where(data_completeInfo[:,ID_index]==NAID_count[i])
    data_completeInfo = np.delete(data_completeInfo,NAID_index,axis=0)

df  = pd.DataFrame(data=data_completeInfo[1:,:],index=data_completeInfo[1:,0],columns=data_completeInfo[0,:])
df2 = df.drop_duplicates(subset='UniqueComplaintId',keep='first')
occ_count = Counter(df2['Borough of Occurrence']).most_common(1)
CompleteInfo_count = len(Counter(data_completeInfo[1:,ID_index]))
occ_count[0][1]/df2.shape[0]

--------------------------------------------------------------

occ_counts = Counter(df2['Borough of Occurrence'])
bor = ['Manhattan','Brooklyn','Queens','Bronx','Staten Island']
pop = {'Manhattan':1643734,'Brooklyn':2629150,'Queens':2333054,'Bronx':1455720,'Staten Island':476015}
ratio=[]
for i in range(5):
    ratio.append(occ_counts[bor[i]]/pop[bor[i]])
max_ind = np.array(np.where(ratio==np.max(ratio)))
com_counts = occ_counts[bor[max_ind[0][0]]]*100000/pop[bor[max_ind[0][0]]]


--------------------------------------------------------------
#Project:
from matplotlib import pyplot as plt
x    = np.loadtxt("wmap_tt_spectra_9yr_v5/wmap_tt_spectrum_9yr_v5.txt")
plt.figure()
plt.plot(x[:,0],x[:,4])
plt.plot(x[:,0],x[:,3])
plt.xlabel('scale')
plt.ylabel('noise')
plt.show()

plt.figure()
plt.plot(x[:,0],x[:,2])
plt.xlabel('scale')
plt.ylabel('error')
plt.show()





